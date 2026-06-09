"""
Saves lead list to a styled Excel file on the Desktop.
Rows are color-coded by priorytet: GORACY=green, CIEPLY=yellow.
"""
import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_desktop = os.path.join(os.path.expanduser("~"), "Desktop")
DESKTOP  = _desktop if os.path.isdir(_desktop) else os.path.expanduser("~")

COLUMNS = [
    ("Priorytet",      "priorytet",  12),
    ("Powod",          "powod",      26),
    ("Nazwa firmy",    "name",       38),
    ("Kategoria",      "category",   20),
    ("Telefon",        "phone",      16),
    ("Email",          "email",      30),
    ("Strona www",     "website",    34),
    ("Adres",          "address",    36),
    ("Opinii",         "reviews",     8),
    ("Ocena",          "rating",      7),
    ("Tel. (www)",     "phone_site", 16),
]

HEADER_BG = "0A1628"
HEADER_FG = "FFFFFF"

FILL_HOT     = PatternFill("solid", fgColor="C6EFCE")   # green  — GORACY
FILL_WARM    = PatternFill("solid", fgColor="FFEB9C")   # yellow — CIEPLY
FILL_DEFAULT = PatternFill("solid", fgColor="F5F5F5")   # light grey — fallback


def save_to_excel(leads: list[dict], path: str | None = None) -> str:
    if not path:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        path = os.path.join(DESKTOP, f"leady_most_digital_{ts}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leady"

    # ── headers ──────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color=HEADER_FG, size=11, name="Calibri")
    hdr_fill  = PatternFill("solid", fgColor=HEADER_BG)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin      = Side(style="thin", color="C0C8D8")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, (header, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 24

    # ── data rows ─────────────────────────────────────────────────────────────
    data_align = Alignment(vertical="center", wrap_text=False)

    for ri, lead in enumerate(leads, 2):
        priorytet = lead.get("priorytet", "")
        if priorytet == "GORACY":
            row_fill = FILL_HOT
        elif priorytet == "CIEPLY":
            row_fill = FILL_WARM
        else:
            row_fill = FILL_DEFAULT

        for ci, (_, key, _) in enumerate(COLUMNS, 1):
            val  = lead.get(key, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = data_align
            cell.font      = Font(size=10, name="Calibri")

            if key == "website" and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(size=10, name="Calibri", color="0563C1", underline="single")
            if key == "email" and val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(size=10, name="Calibri", color="0563C1", underline="single")
            if key == "priorytet":
                cell.font = Font(size=10, name="Calibri", bold=True)

        ws.row_dimensions[ri].height = 17

    # ── summary ───────────────────────────────────────────────────────────────
    last = len(leads) + 2
    hot  = sum(1 for l in leads if l.get("priorytet") == "GORACY")
    warm = sum(1 for l in leads if l.get("priorytet") == "CIEPLY")
    ws.cell(row=last, column=1,
            value=f"Razem: {len(leads)} | Goracy: {hot} | Cieply: {warm}")
    ws.cell(row=last, column=1).font = Font(bold=True, size=10, name="Calibri")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return path
